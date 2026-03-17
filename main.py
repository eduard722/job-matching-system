import requests
import pandas as pd
import time
import re
import urllib3
import json

from sentence_transformers import SentenceTransformer, util
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

model = SentenceTransformer('all-MiniLM-L6-v2')
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

LAST_RESULTS = []

LAST_UPDATE_ID = 0


TELEGRAM_TOKEN = "YOUR_TG_TOKEN"
TELEGRAM_CHAT_ID = "YOUR_CHAT_ID"

SEARCH_QUERY = "data analyst OR аналитик данных"
PER_PAGE = 20
PAGES = 3
TOP_SCORE = 2.5
MID_SCORE = 1.2

SKILLS = {
    "python": 1.5,
    "sql": 1.7,
    "pandas": 1.3,
    "numpy": 0.8,
    "excel": 0.7,
    "git": 0.6,
    "github": 0.5,
    "tableau": 0.5,
    "superset": 0.5,
    "datalens": 0.5,
    "a b": 1.2,
    "ab test": 1.2,
    "статистика": 1.2,
    "метрики": 1.0,
}

GOOD_TITLES = [
    "data analyst",
    "аналитик данных",
    "product analyst",
    "продуктовый аналитик",
    "bi analyst",
    "bi аналитик",
    "web analyst",
    "web аналитик",
    "crm analyst",
    "crm аналитик"
]

try:
    with open("sent.json", "r") as f:
        sent_urls = set(json.load(f))
except:
    sent_urls = set()

try:
    with open("feedback.json", "r") as f:
        FEEDBACK = json.load(f)
except:
    FEEDBACK = {"liked": []}

def clean(text):
    return re.sub(r'\W+', ' ', text.lower())

def skill_score(text):
    return sum(w for skill, w in SKILLS.items() if skill in text)

def title_score(title):
    t = clean(title)
    return 2 if any(x in t for x in GOOD_TITLES) else 0

def bonus_score(text):
    b = 0
    if "intern" in text or "стаж" in text:
        b += 1.5
    if "junior" in text:
        b += 1
    return b

def penalty_score(text):
    t = text.lower()
    p = 0
    if "1 год" in t or "1-3" in t or "2 года" in t:
        p += 0.5
    if "middle" in text:
        p += 0.5
    if "senior" in text:
        p += 3
    if "3 года" in text:
        p += 2
    return p

def stack_bonus(text):
    b = 0
    if "python" in text and "sql" in text:
        b += 1.5
    return b

def explain(text):
    reasons = []
    for s in SKILLS:
        if s in text:
            reasons.append(s)
    return ", ".join(reasons)

def get_vacancies():
    
    vac = []

    for page in range(PAGES):
        r = requests.get(
            "https://api.hh.ru/vacancies",
            params={
                "text": SEARCH_QUERY,
                "per_page": PER_PAGE,
                "page": page,
                "order_by": "publication_time",
                "period": 1
            },
            verify=False
        )

        for item in r.json().get("items", []):
            snippet = item.get("snippet") or {}

            text = (
                (snippet.get("requirement") or "") +
                (snippet.get("responsibility") or "")
            )

            if not text.strip():
                continue

            vac.append({
                "title": item["name"],
                "url": item["alternate_url"],
                "text": text
            })

    return vac

def process(vacancies):
    res = []

    for v in vacancies:
        text = clean(v["text"])

        s = skill_score(text)
        t = title_score(v["title"])
        b = bonus_score(text)
        p = penalty_score(text)
        sb = stack_bonus(text)

        score = s + t + b + sb - p

        if v["url"] in FEEDBACK["liked"]:
            score += 1.5

        reason = explain(text)

        print(v["title"], score)

        if score >= MID_SCORE:
            v["score"] = round(score, 2)
            v["reason"] = reason
            res.append(v)

    return sorted(res, key=lambda x: x["score"], reverse=True)

def save(results):
    wb = Workbook()
    ws = wb.active

    ws.append(["Название", "Score", "Причины", "Ссылка"])

    green = PatternFill(start_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", fill_type="solid")

    for r in results:
        ws.append([r["title"], r["score"], r["reason"], r["url"]])
        row = ws.max_row

        fill = green if r["score"] >= 5 else yellow

        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = fill

    wb.save("hh_results.xlsx")
    print("✅ Excel готов")

def send(results):
    global sent_urls

    new = [r for r in results if r["url"] not in sent_urls]

    print("🗽 Новые вакансии:", len(new))

    if not new:
        print("📭 Нет новых")
        return

    top = []
    mid = []

    for r in new[:10]:
        if r["score"] >= TOP_SCORE:
            top.append(r)
        elif r["score"] >= MID_SCORE:
            mid.append(r)

    for r in top[:10]:
        data = {
            "chat_id": TELEGRAM_CHAT_ID,
            "text": f"🔥 ТОП\n{r['title']}\n📊 {r['score']}",
            "reply_markup": {
                "inline_keyboard": [
                    [
                        {"text": "Открыть", "url": r["url"]}
                    ],
                    [
                        {"text": "👍 Интересно", "callback_data": r["url"]}
                    ]
                ]
            }
        }

        try:
            resp = requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                json=data,
                timeout=5
            )
            print("✅ Отправлено:", resp.status_code)

        except Exception as e:
            print("Telegram ошибка:", e)


        sent_urls.add(r["url"])
        
        time.sleep(1)

    for r in mid[:10]:
        data = {
            "chat_id": TELEGRAM_CHAT_ID,
            "text": f"🟡 Можно рассмотреть\n{r['title']}\n📊 {r['score']}",
            "reply_markup": {
                "inline_keyboard": [[
                    {"text": "Открыть", "url": r["url"]}
                ]]
            }
        }

        try:
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                json=data,
                timeout=5
            )
        except Exception as e:
            print("Telegram ошибка:", e)


        sent_urls.add(r["url"])

    with open("sent.json", "w") as f:
        json.dump(list(sent_urls), f)

def main():
    while True:
        print("1️⃣ feedback")
        handle_feedback()

        print("🔄 Поиск...")        
        vac = get_vacancies()

        print("3️⃣ Обработка")
        res = process(vac)

        global LAST_RESULTS
        LAST_RESULTS = res

        print("4️⃣ Сохранение")
        save(res)
        print("📤 Отправка в Telegram...")
        send(res)
        print("✅ Telegram отправлен")

        print(f"✅ Готово | Всего: {len(vac)} | Подошло: {len(res)}")
        print("⏳ Ждём 10 минут\n")

        for _ in range(120):
            handle_feedback()
            time.sleep(0.5)

def handle_feedback():
    global LAST_RESULTS
    global LAST_UPDATE_ID

    try:
        updates = requests.get(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates",
            params={
                "offset": LAST_UPDATE_ID + 1,
                "timeout":10
                },
            timeout=15
        ).json()

        for u in updates.get("result", []):
            LAST_UPDATE_ID = u["update_id"]

            if "callback_query" in u:
                callback_id = u["callback_query"]["id"]
                url = u["callback_query"]["data"]

                requests.post(
                    f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/answerCallbackQuery",
                    json={
                        "callback_query_id": callback_id,
                        "text": "Сохранено 🍀",
                        "show_alert": False
                    },
                    timeout=2
                )

                for v in LAST_RESULTS:
                    if v["url"] == url:
                        liked_text = v["text"]

                        FEEDBACK["liked"].append(liked_text)

                        with open("feedback.json", "w") as f:
                            json.dump(FEEDBACK, f)

                        print("👍 Лайк сохранен:", url)
                        break

    except Exception as e:
        print("Ошибка feedback:", e)

def semantic_score(text):
    if not FEEDBACK["liked"]:
        return 0

    try:
        current_embedding = model.encode(text, convert_to_tensor=True)

        max_sim = 0

        for liked_text in FEEDBACK["liked"]:
            liked_embedding = model.encode(liked_text, convert_to_tensor=True)

            sim = util.cos_sim(current_embedding, liked_embedding).item()

            if sim > max_sim:
                max_sim = sim

        return max_sim * 2

    except:
        return 0

if __name__ == "__main__":
    main()